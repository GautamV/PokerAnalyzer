import os
import sys
import csv
import re

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

PLAYER_STACKS_MATCHER = re.compile(r'(?<=Player stacks:).*')
NUMBER_MATCHER = re.compile(r'(?<=#)\d+(?= )')
NAME_MATCHER = re.compile(r'(?<=").+(?= @)')
STACK_MATCHER = re.compile(r'(?<=\().+(?=\))') 

YOUR_HAND_MATCHER = re.compile(r'(?<=Your hand is ).*')
SHOWS_HAND_MATCHER = re.compile(r'(?<=shows a ).*')

SMALL_BLIND_MATCHER = re.compile(r'(?<=posts a small blind of )\d+(?=\s|$)')
BIG_BLIND_MATCHER = re.compile(r'(?<=posts a big blind of )\d+(?=\s|$)')

BETS_MATCHER = re.compile(r'(?<=bets )\d+(?=\s|$)')
RAISES_MATCHER = re.compile(r'(?<=raises to )\d+(?=\s|$)')
CALLS_MATCHER = re.compile(r'(?<=calls )\d+(?=\s|$)')

DARK_BLUE_FILL = PatternFill(fgColor='2EBCDB', fill_type = "solid")
LIGHT_BLUE_FILL = PatternFill(fgColor='99ECFF', fill_type = "solid")
LIGHT_RED_FILL = PatternFill(fgColor='FA8C8C', fill_type = "solid")
LIGHT_GREEN_FILL = PatternFill(fgColor='66FF7F', fill_type = "solid")
YELLOW_FILL = PatternFill(fgColor='F4FA8C', fill_type = "solid")

COLUMN_INDECES = {
	'Action': 0, 
	'Amount': 1, 
	'Stack': 2, 
	'Pot': 3,
	'Time': 4,
	'Name': 0,
	'Hand': 3, 
}

def get_column(action, player_num): 
	player_num = int(player_num)
	return 3 + (5 * (player_num - 1)) + COLUMN_INDECES[action]

def get_time(timestamp): 
	return datetime.strptime(timestamp, '%Y-%m-%dT%H:%M:%S.%fZ')

class Analyzer: 
	def __init__(self, filepath, player_name): 
		self.player_name = player_name
		self.player_wins = {}

		self.init_sheet()

		csv_reader = csv.reader(reversed(open(filepath, newline='', encoding='utf8').readlines()), delimiter=',', quotechar='"')
		for line in csv_reader: 
			self.process_line(line)

		self.build_summary_sheet()

	def save_sheet(self, filepath):
		self.wb.save(filename=filepath)

	def init_sheet(self):
		self.wb = Workbook()
		self.summary = self.wb.active
		self.summary.title = 'Summary'
		self.all_hands = self.wb.create_sheet('All Hands')
		self.row = 1
		self.last_player_name_row = -1

		self.hand_counter = 1

		self.all_hands.cell(row=self.row, column=1, value='Hands')
		self.all_hands.cell(row=self.row, column=2, value='Phase')
		for i in range(10): 
			player_group_start = (3+(i*5))
			self.all_hands.cell(self.row, column=player_group_start, value='Player {}'.format(i+1))
			self.all_hands.merge_cells(start_row=self.row, end_row=self.row, start_column=player_group_start, end_column=player_group_start+4)

		bold_font = Font(bold=True)
		gray_fill = PatternFill(fgColor="7D7D7D", fill_type = "solid")
		for cell in self.all_hands["1:1"]:
		    cell.font = bold_font
		    cell.fill = gray_fill

		self.all_hands.column_dimensions['B'].width = 40

	def process_line(self, line): 
		message, timestamp, _ = line
		if 'Player stacks:' in message:
			self.process_player_stacks(message)
			self.last_action_time = get_time(timestamp)
		elif 'Your hand is' in message:
			self.process_your_hand(message)
			self.last_action_time = get_time(timestamp)
		elif 'posts a small blind' in message:
			self.process_small_blind(message)
			self.last_action_time = get_time(timestamp)
		elif 'posts a big blind' in message:
			self.process_big_blind(message)
			self.last_action_time = get_time(timestamp)
		elif 'raises' in message:
			self.process_raise(message, timestamp)
			self.last_action_time = get_time(timestamp)
		elif 'checks' in message:
			self.process_check(message, timestamp)
			self.last_action_time = get_time(timestamp)
		elif 'bets' in message:	
			self.process_bet(message, timestamp)
			self.last_action_time = get_time(timestamp)
		elif 'folds' in message:
			self.process_fold(message, timestamp)
			self.last_action_time = get_time(timestamp)
		elif 'calls' in message:
			self.process_call(message, timestamp)
			self.last_action_time = get_time(timestamp)
		elif 'Flop:' in message or 'Turn:' in message or 'River:' in message:
			self.process_phase(message)
			self.last_action_time = get_time(timestamp)
		elif 'shows a' in message:
			self.process_shows_hand(message)
			self.last_action_time = get_time(timestamp)
		elif 'collected' in message:
			self.process_win(message)
			self.last_action_time = get_time(timestamp)

	def process_player_stacks(self, message): 
		self.row += 1
		player_list = PLAYER_STACKS_MATCHER.search(message).group().split('|')
		self.player_name_to_number = {}
		self.player_bet_amounts = {}
		self.pot = 0
		self.last_bet_or_raise = 0

		self.all_hands.cell(row=self.row, column=1, value=self.hand_counter)
		self.hand_counter += 1 

		for player in player_list: 
			number = NUMBER_MATCHER.search(player).group()
			name = NAME_MATCHER.search(player).group()
			stack = STACK_MATCHER.search(player).group()

			self.player_name_to_number[name] = int(number)

			self.all_hands.cell(row=self.row, column=get_column('Name', number), value=name)
			self.all_hands.merge_cells(start_row=self.row, end_row=self.row, start_column=get_column('Name', number), end_column=get_column('Name', number)+2)
			self.all_hands.merge_cells(start_row=self.row, end_row=self.row, start_column=get_column('Hand', number), end_column=get_column('Hand', number)+1)

			self.all_hands.cell(row=self.row+1, column=get_column('Action', number), value='Action')
			self.all_hands.cell(row=self.row+1, column=get_column('Amount', number), value='Amount')
			self.all_hands.cell(row=self.row+1, column=get_column('Stack', number), value='Stack')
			self.all_hands.cell(row=self.row+1, column=get_column('Pot', number), value='Pot')
			self.all_hands.cell(row=self.row+1, column=get_column('Time', number), value='Time')

			self.all_hands.cell(row=self.row+2, column=get_column('Stack', number), value=stack)

		for cell in self.all_hands['{}:{}'.format(self.row, self.row)]: 
			cell.fill = DARK_BLUE_FILL

		self.last_player_name_row = self.row
		self.row += 2

	def process_your_hand(self, message): 
		if self.player_name not in self.player_name_to_number: 
			return
		hand = YOUR_HAND_MATCHER.search(message).group()
		self.all_hands.cell(row=self.last_player_name_row, column=get_column('Hand', self.player_name_to_number[self.player_name]), value=hand)

	def process_small_blind(self, message):
		self.row += 1
		name = NAME_MATCHER.search(message).group()
		value = int(SMALL_BLIND_MATCHER.search(message).group())

		amount = value - (self.player_bet_amounts[name] if name in self.player_bet_amounts else 0)
		stack = self.find_previous_stack(self.row, name) - amount
		self.player_bet_amounts[name] = value
		self.last_bet_or_raise = value

		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='SB')
		self.all_hands.cell(row=self.row, column=get_column('Amount', self.player_name_to_number[name]), value=amount)
		self.all_hands.cell(row=self.row, column=get_column('Stack', self.player_name_to_number[name]), value=stack)
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())
		for col in range(get_column('Action', self.player_name_to_number[name]), get_column('Action', self.player_name_to_number[name]) + 5): 
			self.all_hands.cell(row=self.row, column=col).fill = YELLOW_FILL

	def process_big_blind(self, message):
		self.row += 1
		name = NAME_MATCHER.search(message).group()
		value = int(BIG_BLIND_MATCHER.search(message).group())

		amount = value - (self.player_bet_amounts[name] if name in self.player_bet_amounts else 0)
		stack = self.find_previous_stack(self.row, name) - amount
		self.player_bet_amounts[name] = value
		self.last_bet_or_raise = value

		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='BB')
		self.all_hands.cell(row=self.row, column=get_column('Amount', self.player_name_to_number[name]), value=amount)
		self.all_hands.cell(row=self.row, column=get_column('Stack', self.player_name_to_number[name]), value=stack)
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())
		for col in range(get_column('Action', self.player_name_to_number[name]), get_column('Action', self.player_name_to_number[name]) + 5): 
			self.all_hands.cell(row=self.row, column=col).fill = YELLOW_FILL

	def process_shows_hand(self, message):
		hand = SHOWS_HAND_MATCHER.search(message).group()
		name = NAME_MATCHER.search(message).group()
		self.all_hands.cell(row=self.last_player_name_row, column=get_column('Hand', self.player_name_to_number[name]), value=hand)

	def process_fold(self, message, timestamp): 
		name = NAME_MATCHER.search(message).group()
		seconds = int((get_time(timestamp) - self.last_action_time).total_seconds())

		stack = self.find_previous_stack(self.row, name)
		amount = self.last_bet_or_raise - (self.player_bet_amounts[name] if name in self.player_bet_amounts else 0)

		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='Fold')
		self.all_hands.cell(row=self.row, column=get_column('Amount', self.player_name_to_number[name]), value='({})'.format(amount))
		self.all_hands.cell(row=self.row, column=get_column('Time', self.player_name_to_number[name]), value='{}s'.format(seconds))
		self.all_hands.cell(row=self.row, column=get_column('Stack', self.player_name_to_number[name]), value=stack)
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())
		for col in range(get_column('Action', self.player_name_to_number[name]), get_column('Action', self.player_name_to_number[name]) + 5): 
			self.all_hands.cell(row=self.row, column=col).fill = LIGHT_RED_FILL

	def process_check(self, message, timestamp): 
		name = NAME_MATCHER.search(message).group()
		seconds = int((get_time(timestamp) - self.last_action_time).total_seconds())

		stack = self.find_previous_stack(self.row, name)

		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='Check')
		self.all_hands.cell(row=self.row, column=get_column('Amount', self.player_name_to_number[name]), value='0')
		self.all_hands.cell(row=self.row, column=get_column('Time', self.player_name_to_number[name]), value='{}s'.format(seconds))
		self.all_hands.cell(row=self.row, column=get_column('Stack', self.player_name_to_number[name]), value=stack)
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())

	def process_bet(self, message, timestamp): 
		self.row += 1
		name = NAME_MATCHER.search(message).group()
		value = int(BETS_MATCHER.search(message).group())
		seconds = int((get_time(timestamp) - self.last_action_time).total_seconds())

		amount = value - (self.player_bet_amounts[name] if name in self.player_bet_amounts else 0)
		stack = self.find_previous_stack(self.row, name) - amount
		self.player_bet_amounts[name] = value
		self.last_bet_or_raise = value

		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='Bet')
		self.all_hands.cell(row=self.row, column=get_column('Amount', self.player_name_to_number[name]), value=amount)
		self.all_hands.cell(row=self.row, column=get_column('Time', self.player_name_to_number[name]), value='{}s'.format(seconds))
		self.all_hands.cell(row=self.row, column=get_column('Stack', self.player_name_to_number[name]), value=stack)
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())
		for col in range(get_column('Action', self.player_name_to_number[name]), get_column('Action', self.player_name_to_number[name]) + 5): 
			self.all_hands.cell(row=self.row, column=col).fill = YELLOW_FILL

	def process_call(self, message, timestamp): 
		name = NAME_MATCHER.search(message).group()
		value = int(CALLS_MATCHER.search(message).group())
		seconds = int((get_time(timestamp) - self.last_action_time).total_seconds())

		amount = value - (self.player_bet_amounts[name] if name in self.player_bet_amounts else 0)
		stack = self.find_previous_stack(self.row, name) - amount
		self.player_bet_amounts[name] = value

		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='Call')
		self.all_hands.cell(row=self.row, column=get_column('Amount', self.player_name_to_number[name]), value=amount)
		self.all_hands.cell(row=self.row, column=get_column('Time', self.player_name_to_number[name]), value='{}s'.format(seconds))
		self.all_hands.cell(row=self.row, column=get_column('Stack', self.player_name_to_number[name]), value=stack)
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())

	def process_raise(self, message, timestamp): 
		self.row += 1 
		name = NAME_MATCHER.search(message).group()
		value = int(RAISES_MATCHER.search(message).group())
		seconds = int((get_time(timestamp) - self.last_action_time).total_seconds())

		amount = value - (self.player_bet_amounts[name] if name in self.player_bet_amounts else 0)
		stack = self.find_previous_stack(self.row, name) - amount
		self.player_bet_amounts[name] = value
		self.last_bet_or_raise = value

		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='Raise')
		self.all_hands.cell(row=self.row, column=get_column('Amount', self.player_name_to_number[name]), value=amount)
		self.all_hands.cell(row=self.row, column=get_column('Time', self.player_name_to_number[name]), value='{}s'.format(seconds))
		self.all_hands.cell(row=self.row, column=get_column('Stack', self.player_name_to_number[name]), value=stack)
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())
		for col in range(get_column('Action', self.player_name_to_number[name]), get_column('Action', self.player_name_to_number[name]) + 5): 
			self.all_hands.cell(row=self.row, column=col).fill = YELLOW_FILL

	def process_phase(self, message):
		self.row += 1 
		self.all_hands.cell(row=self.row, column=2, value=message).fill = LIGHT_BLUE_FILL
		self.pot += sum(self.player_bet_amounts.values())
		self.player_bet_amounts = {}
		self.last_bet_or_raise = 0

	def process_win(self, message): 
		self.row += 1
		name = NAME_MATCHER.search(message).group()
		self.all_hands.cell(row=self.row, column=get_column('Action', self.player_name_to_number[name]), value='Win')
		self.all_hands.cell(row=self.row, column=get_column('Pot', self.player_name_to_number[name]), value=self.calculate_pot())
		for col in range(get_column('Action', self.player_name_to_number[name]), get_column('Action', self.player_name_to_number[name]) + 5): 
			self.all_hands.cell(row=self.row, column=col).fill = LIGHT_GREEN_FILL

		if name in self.player_wins: 
			self.player_wins[name].append(self.calculate_pot())
		else: 
			self.player_wins[name] = [self.calculate_pot()]

	def find_previous_stack(self, current_row, player_name): 
		row = current_row - 1
		while self.all_hands.cell(row=row, column=get_column('Stack', self.player_name_to_number[player_name])).value is None: 
			row -= 1 
		return int(self.all_hands.cell(row=row, column=get_column('Stack', self.player_name_to_number[player_name])).value)

	def calculate_pot(self): 
		return self.pot + sum(self.player_bet_amounts.values())

	def build_summary_sheet(self):
		self.summary.column_dimensions['A'].width = 40
		self.summary['A1'] = 'Player:'
		self.summary['A2'] = 'Number of Wins:'
		self.summary['A3'] = 'Average Win Size:'

		col = 2
		for name, wins in self.player_wins.items(): 
			self.summary.cell(row=1, column=col, value=name)
			self.summary.cell(row=2, column=col, value=len(wins))
			self.summary.cell(row=3, column=col, value=(sum(wins)/len(wins)))
			col += 1 



### FOR TESTING ### 

OUTPUT_FILEPATH = 'poker_log_analysis.xlsx'

def run_from_terminal(): 
	if len(sys.argv) != 3: 
		print('ERROR: Must pass in exactly two arguments (path to csv poker log, your name)')
		return

	analyzer = Analyzer(os.path.join(sys.path[0], sys.argv[1]), sys.argv[2])
	analyzer.save_sheet(OUTPUT_FILEPATH)
	
if __name__ == '__main__': 
	run_from_terminal()